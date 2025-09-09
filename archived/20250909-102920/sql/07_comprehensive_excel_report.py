#!/usr/bin/env python3
"""
Comprehensive Excel Template Analysis Report
This generates a detailed report showing ALL matched Excel templates with:
- Original complexity indicators
- Derived complexity metrics
- Field categorization (unique vs shared)
- Sharing patterns and template relationships
"""

import psycopg2
import pandas as pd
from datetime import datetime
import numpy as np
from pathlib import Path
import logging
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils.dataframe import dataframe_to_rows

logging.basicConfig(level=logging.INFO, format='%(asctime)s - %(levelname)s - %(message)s')
logger = logging.getLogger(__name__)

# Database configuration
DB_CONFIG = {
    'host': 'localhost',
    'port': 5432,
    'database': 'estimatedoc',
    'user': 'estimatedoc_user',
    'password': 'estimatedoc_user'
}

class ExcelTemplateReporter:
    def __init__(self):
        self.conn = psycopg2.connect(**DB_CONFIG)
        self.cursor = self.conn.cursor()
        
    def __del__(self):
        if hasattr(self, 'conn'):
            self.conn.close()
    
    def analyze_requirements(self):
        """
        Deep thinking about report structure:
        
        For each Excel template row, we need:
        1. Identity: Row ID, Template Code, Description
        2. Original Assessment: Stated Complexity from Excel
        3. Match Status: How we matched it, confidence level
        4. Actual Metrics: 
           - Total fields count
           - Unique fields (only in this doc)
           - Shared fields (in multiple docs)
           - Sharing breadth (how many other docs share)
        5. Field Categories:
           - DOCVARIABLE vs MERGEFIELD
           - IF statements (conditional logic)
           - Nested IF statements (complex logic)
           - User input fields
           - Calculated fields
        6. Reusability Score: Percentage of fields that are shared
        7. Complexity Validation: Does actual match stated?
        """
        logger.info("Report structure analyzed - proceeding with comprehensive design")
    
    def get_matched_templates_with_full_metrics(self):
        """
        Get ALL matched Excel templates with comprehensive metrics
        """
        query = """
        WITH template_field_details AS (
            -- Get detailed field information for each matched template
            SELECT 
                tm.excel_row_id,
                tm.document_id,
                f.field_id,
                f.field_code,
                -- Count how many documents use this field
                (SELECT COUNT(DISTINCT document_id) 
                 FROM document_fields df2 
                 WHERE df2.field_id = f.field_id) as field_usage_count,
                -- Get list of other templates that share this field
                (SELECT COUNT(DISTINCT tm2.excel_row_id)
                 FROM document_fields df2
                 JOIN template_matches tm2 ON df2.document_id = tm2.document_id
                 WHERE df2.field_id = f.field_id 
                   AND tm2.excel_row_id != tm.excel_row_id) as shared_with_templates_count
            FROM template_matches tm
            JOIN document_fields df ON tm.document_id = df.document_id
            JOIN fields f ON df.field_id = f.field_id
            WHERE tm.document_id IS NOT NULL
        ),
        field_aggregates AS (
            -- Aggregate field metrics per template
            SELECT 
                excel_row_id,
                document_id,
                COUNT(DISTINCT field_id) as total_fields,
                COUNT(DISTINCT CASE WHEN field_usage_count = 1 THEN field_id END) as unique_fields,
                COUNT(DISTINCT CASE WHEN field_usage_count > 1 THEN field_id END) as shared_fields,
                COUNT(DISTINCT CASE WHEN shared_with_templates_count > 0 THEN field_id END) as shared_with_other_templates,
                ROUND(AVG(shared_with_templates_count), 2) as avg_sharing_breadth,
                MAX(shared_with_templates_count) as max_sharing_breadth,
                -- Field categorization
                COUNT(DISTINCT CASE WHEN field_code LIKE '%DOCVARIABLE%' THEN field_id END) as docvariable_count,
                COUNT(DISTINCT CASE WHEN field_code LIKE '%MERGEFIELD%' THEN field_id END) as mergefield_count,
                COUNT(DISTINCT CASE WHEN field_code LIKE '%IF %' THEN field_id END) as if_statement_fields,
                COUNT(DISTINCT CASE WHEN field_code LIKE '%ASK%' OR field_code LIKE '%FILLIN%' THEN field_id END) as user_input_fields,
                COUNT(DISTINCT CASE WHEN field_code LIKE '%=%' OR field_code LIKE '%FORMULA%' THEN field_id END) as calculated_fields
            FROM template_field_details
            GROUP BY excel_row_id, document_id
        )
        SELECT 
            et.row_id,
            et.template_code,
            et.description,
            et.stated_complexity,
            -- Matching details
            tm.match_method,
            tm.match_confidence,
            tm.manifest_code,
            tm.manifest_name,
            -- Document details
            d.filename as document_filename,
            d.basename as document_basename,
            -- Field metrics from aggregates
            COALESCE(fa.total_fields, 0) as total_fields,
            COALESCE(fa.unique_fields, 0) as unique_to_this_doc,
            COALESCE(fa.shared_fields, 0) as shared_with_others,
            COALESCE(fa.shared_with_other_templates, 0) as templates_sharing_fields,
            COALESCE(fa.avg_sharing_breadth, 0) as avg_templates_per_shared_field,
            COALESCE(fa.max_sharing_breadth, 0) as max_templates_sharing_single_field,
            -- Reusability percentage
            CASE 
                WHEN fa.total_fields > 0 
                THEN ROUND(100.0 * fa.shared_fields / fa.total_fields, 1)
                ELSE 0
            END as reusability_percentage,
            -- Field categories
            COALESCE(fa.docvariable_count, 0) as docvariable_fields,
            COALESCE(fa.mergefield_count, 0) as mergefield_fields,
            COALESCE(fa.if_statement_fields, 0) as conditional_logic_fields,
            COALESCE(fa.user_input_fields, 0) as user_input_fields,
            COALESCE(fa.calculated_fields, 0) as calculated_fields,
            -- Complexity metrics from document_complexity table
            COALESCE(dc.if_statements, 0) as if_statements_count,
            COALESCE(dc.nested_if_statements, 0) as nested_if_count,
            COALESCE(dc.total_complexity_score, 0) as calculated_complexity_score,
            COALESCE(dc.complexity_rating, 'Unknown') as calculated_complexity_rating,
            -- Complexity validation
            CASE 
                WHEN et.stated_complexity = 'Simple' AND COALESCE(dc.total_complexity_score, 0) <= 10 THEN 'Matches'
                WHEN et.stated_complexity = 'Moderate' AND COALESCE(dc.total_complexity_score, 0) BETWEEN 11 AND 25 THEN 'Matches'
                WHEN et.stated_complexity = 'Complex' AND COALESCE(dc.total_complexity_score, 0) > 25 THEN 'Matches'
                WHEN COALESCE(dc.total_complexity_score, 0) = 0 THEN 'No Data'
                ELSE 'Mismatch'
            END as complexity_validation,
            -- Additional insights
            CASE 
                WHEN fa.unique_fields > fa.shared_fields THEN 'Mostly Unique'
                WHEN fa.shared_fields > fa.unique_fields * 2 THEN 'Highly Reusable'
                WHEN fa.shared_fields > 0 THEN 'Balanced'
                ELSE 'No Fields'
            END as reusability_category
        FROM excel_templates et
        JOIN template_matches tm ON et.row_id = tm.excel_row_id
        LEFT JOIN documents d ON tm.document_id = d.document_id
        LEFT JOIN field_aggregates fa ON tm.excel_row_id = fa.excel_row_id
        LEFT JOIN document_complexity dc ON tm.document_id = dc.document_id
        WHERE tm.document_id IS NOT NULL
        ORDER BY et.row_id
        """
        
        logger.info("Executing comprehensive template analysis query...")
        df = pd.read_sql_query(query, self.conn)
        logger.info(f"Retrieved {len(df)} matched templates with full metrics")
        return df
    
    def get_sharing_matrix(self):
        """
        Create a matrix showing which templates share fields with each other
        """
        query = """
        WITH template_pairs AS (
            SELECT DISTINCT
                tm1.excel_row_id as template1_id,
                et1.template_code as template1_code,
                tm2.excel_row_id as template2_id,
                et2.template_code as template2_code,
                COUNT(DISTINCT df1.field_id) as shared_fields_count
            FROM template_matches tm1
            JOIN document_fields df1 ON tm1.document_id = df1.document_id
            JOIN document_fields df2 ON df1.field_id = df2.field_id
            JOIN template_matches tm2 ON df2.document_id = tm2.document_id
            JOIN excel_templates et1 ON tm1.excel_row_id = et1.row_id
            JOIN excel_templates et2 ON tm2.excel_row_id = et2.row_id
            WHERE tm1.excel_row_id < tm2.excel_row_id  -- Avoid duplicates and self-joins
              AND tm1.document_id IS NOT NULL
              AND tm2.document_id IS NOT NULL
            GROUP BY tm1.excel_row_id, et1.template_code, tm2.excel_row_id, et2.template_code
            HAVING COUNT(DISTINCT df1.field_id) > 5  -- Only show significant sharing
        )
        SELECT * FROM template_pairs
        ORDER BY shared_fields_count DESC
        LIMIT 100  -- Top 100 sharing relationships
        """
        
        logger.info("Analyzing template sharing relationships...")
        df = pd.read_sql_query(query, self.conn)
        logger.info(f"Found {len(df)} significant template sharing relationships")
        return df
    
    def get_complexity_comparison(self):
        """
        Compare stated vs calculated complexity
        """
        query = """
        WITH complexity_analysis AS (
            SELECT 
                et.stated_complexity,
                COUNT(*) as template_count,
                AVG(COALESCE(dc.total_complexity_score, 0)) as avg_calculated_score,
                MIN(COALESCE(dc.total_complexity_score, 0)) as min_calculated_score,
                MAX(COALESCE(dc.total_complexity_score, 0)) as max_calculated_score,
                STDDEV(COALESCE(dc.total_complexity_score, 0)) as stddev_calculated_score,
                COUNT(CASE 
                    WHEN et.stated_complexity = 'Simple' AND COALESCE(dc.total_complexity_score, 0) <= 10 THEN 1
                    WHEN et.stated_complexity = 'Moderate' AND COALESCE(dc.total_complexity_score, 0) BETWEEN 11 AND 25 THEN 1
                    WHEN et.stated_complexity = 'Complex' AND COALESCE(dc.total_complexity_score, 0) > 25 THEN 1
                END) as matching_assessments,
                COUNT(CASE 
                    WHEN NOT (
                        (et.stated_complexity = 'Simple' AND COALESCE(dc.total_complexity_score, 0) <= 10) OR
                        (et.stated_complexity = 'Moderate' AND COALESCE(dc.total_complexity_score, 0) BETWEEN 11 AND 25) OR
                        (et.stated_complexity = 'Complex' AND COALESCE(dc.total_complexity_score, 0) > 25)
                    ) AND COALESCE(dc.total_complexity_score, 0) > 0 THEN 1
                END) as mismatched_assessments
            FROM excel_templates et
            JOIN template_matches tm ON et.row_id = tm.excel_row_id
            LEFT JOIN document_complexity dc ON tm.document_id = dc.document_id
            WHERE tm.document_id IS NOT NULL
              AND et.stated_complexity IS NOT NULL
            GROUP BY et.stated_complexity
        )
        SELECT 
            stated_complexity,
            template_count,
            ROUND(avg_calculated_score, 2) as avg_score,
            min_calculated_score as min_score,
            max_calculated_score as max_score,
            ROUND(stddev_calculated_score, 2) as score_variation,
            matching_assessments,
            mismatched_assessments,
            ROUND(100.0 * matching_assessments / NULLIF(template_count, 0), 1) as accuracy_percentage
        FROM complexity_analysis
        ORDER BY 
            CASE stated_complexity
                WHEN 'Simple' THEN 1
                WHEN 'Moderate' THEN 2
                WHEN 'Complex' THEN 3
                ELSE 4
            END
        """
        
        logger.info("Comparing stated vs calculated complexity...")
        df = pd.read_sql_query(query, self.conn)
        return df
    
    def get_field_reuse_leaders(self):
        """
        Identify the most commonly reused fields
        """
        query = """
        WITH field_usage AS (
            SELECT 
                f.field_id,
                REPLACE(REPLACE(REPLACE(f.field_code, CHR(13), ' '), CHR(10), ' '), CHR(9), ' ') as field_code_clean,
                COUNT(DISTINCT df.document_id) as document_count,
                COUNT(DISTINCT tm.excel_row_id) as template_count,
                STRING_AGG(DISTINCT et.template_code, ', ' ORDER BY et.template_code) as used_in_templates
            FROM fields f
            JOIN document_fields df ON f.field_id = df.field_id
            JOIN template_matches tm ON df.document_id = tm.document_id
            JOIN excel_templates et ON tm.excel_row_id = et.row_id
            WHERE tm.document_id IS NOT NULL
            GROUP BY f.field_id, f.field_code
            HAVING COUNT(DISTINCT tm.excel_row_id) > 10  -- Used in more than 10 templates
        )
        SELECT 
            field_code_clean as field_code,
            template_count,
            CASE 
                WHEN LENGTH(used_in_templates) > 100 
                THEN SUBSTRING(used_in_templates, 1, 100) || '...'
                ELSE used_in_templates
            END as sample_templates
        FROM field_usage
        ORDER BY template_count DESC
        LIMIT 50
        """
        
        logger.info("Identifying most reused fields...")
        df = pd.read_sql_query(query, self.conn)
        # Additional cleaning for any remaining special characters
        if not df.empty:
            df['field_code'] = df['field_code'].apply(lambda x: ''.join(c if ord(c) >= 32 else ' ' for c in str(x)))
        return df
    
    def generate_excel_report(self, output_path='Excel_Template_Comprehensive_Analysis.xlsx'):
        """
        Generate comprehensive Excel report with multiple sheets
        """
        logger.info(f"Generating comprehensive Excel report: {output_path}")
        
        # Collect all data
        main_df = self.get_matched_templates_with_full_metrics()
        sharing_df = self.get_sharing_matrix()
        complexity_df = self.get_complexity_comparison()
        reuse_leaders_df = self.get_field_reuse_leaders()
        
        # Create Excel writer with xlsxwriter engine for better formatting
        with pd.ExcelWriter(output_path, engine='openpyxl') as writer:
            # Main comprehensive sheet
            main_df.to_excel(writer, sheet_name='Template Analysis', index=False)
            self._format_main_sheet(writer.book['Template Analysis'], main_df)
            
            # Complexity comparison
            complexity_df.to_excel(writer, sheet_name='Complexity Validation', index=False)
            
            # Sharing matrix
            if not sharing_df.empty:
                sharing_df.to_excel(writer, sheet_name='Template Sharing', index=False)
            
            # Most reused fields
            if not reuse_leaders_df.empty:
                reuse_leaders_df.to_excel(writer, sheet_name='Top Reused Fields', index=False)
            
            # Summary statistics
            summary_df = self._create_summary_statistics(main_df)
            summary_df.to_excel(writer, sheet_name='Summary', index=False)
            
        logger.info(f"Report successfully generated: {output_path}")
        return output_path
    
    def _format_main_sheet(self, worksheet, df):
        """
        Apply formatting to the main analysis sheet
        """
        # Set column widths
        column_widths = {
            'A': 8,   # row_id
            'B': 15,  # template_code
            'C': 50,  # description
            'D': 12,  # stated_complexity
            'E': 20,  # match_method
            'F': 15,  # match_confidence
            'G': 15,  # total_fields
            'H': 18,  # unique_to_this_doc
            'I': 18,  # shared_with_others
            'J': 20,  # reusability_percentage
            'K': 20,  # complexity_validation
        }
        
        for col, width in column_widths.items():
            worksheet.column_dimensions[col].width = width
        
        # Apply header formatting
        header_font = Font(bold=True, color="FFFFFF")
        header_fill = PatternFill(start_color="366092", end_color="366092", fill_type="solid")
        header_alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
        
        for cell in worksheet[1]:
            cell.font = header_font
            cell.fill = header_fill
            cell.alignment = header_alignment
        
        # Conditional formatting for complexity validation
        for row in range(2, len(df) + 2):
            validation_cell = worksheet[f'K{row}']
            if validation_cell.value == 'Matches':
                validation_cell.fill = PatternFill(start_color="C6EFCE", end_color="C6EFCE", fill_type="solid")
            elif validation_cell.value == 'Mismatch':
                validation_cell.fill = PatternFill(start_color="FFC7CE", end_color="FFC7CE", fill_type="solid")
        
        # Freeze panes
        worksheet.freeze_panes = 'A2'
    
    def _create_summary_statistics(self, df):
        """
        Create summary statistics from the main dataframe
        """
        summary = {
            'Metric': [],
            'Value': []
        }
        
        # Overall statistics
        total_templates = len(df)
        avg_fields = df['total_fields'].mean()
        avg_unique = df['unique_to_this_doc'].mean()
        avg_shared = df['shared_with_others'].mean()
        avg_reusability = df['reusability_percentage'].mean()
        
        # Complexity validation
        matches = len(df[df['complexity_validation'] == 'Matches'])
        mismatches = len(df[df['complexity_validation'] == 'Mismatch'])
        
        # High reusability templates
        high_reuse = len(df[df['reusability_percentage'] > 75])
        low_reuse = len(df[df['reusability_percentage'] < 25])
        
        summary['Metric'] = [
            'Total Matched Templates',
            'Average Fields per Template',
            'Average Unique Fields',
            'Average Shared Fields',
            'Average Reusability %',
            '',
            'Complexity Assessment Matches',
            'Complexity Assessment Mismatches',
            'Match Accuracy %',
            '',
            'Highly Reusable Templates (>75%)',
            'Low Reusability Templates (<25%)',
            '',
            'Templates with IF Statements',
            'Templates with User Input Fields',
            'Templates with Calculated Fields'
        ]
        
        summary['Value'] = [
            total_templates,
            round(avg_fields, 1),
            round(avg_unique, 1),
            round(avg_shared, 1),
            round(avg_reusability, 1),
            '',
            matches,
            mismatches,
            round(100 * matches / (matches + mismatches), 1) if (matches + mismatches) > 0 else 0,
            '',
            high_reuse,
            low_reuse,
            '',
            len(df[df['if_statements_count'] > 0]),
            len(df[df['user_input_fields'] > 0]),
            len(df[df['calculated_fields'] > 0])
        ]
        
        return pd.DataFrame(summary)
    
    def print_key_insights(self, df):
        """
        Print key insights from the analysis
        """
        logger.info("\n" + "="*80)
        logger.info("KEY INSIGHTS FROM EXCEL TEMPLATE ANALYSIS")
        logger.info("="*80)
        
        # Overall metrics
        logger.info(f"\n📊 OVERALL METRICS")
        logger.info(f"  Total Matched Templates: {len(df)}")
        logger.info(f"  Average Fields per Template: {df['total_fields'].mean():.1f}")
        logger.info(f"  Average Reusability: {df['reusability_percentage'].mean():.1f}%")
        
        # Complexity validation
        matches = len(df[df['complexity_validation'] == 'Matches'])
        total = len(df[df['complexity_validation'].isin(['Matches', 'Mismatch'])])
        logger.info(f"\n✅ COMPLEXITY VALIDATION")
        logger.info(f"  Accuracy: {100*matches/total:.1f}% ({matches}/{total} templates)")
        
        # Top reusable templates
        top_reusable = df.nlargest(5, 'reusability_percentage')[['template_code', 'reusability_percentage']]
        logger.info(f"\n🔄 TOP REUSABLE TEMPLATES")
        for _, row in top_reusable.iterrows():
            logger.info(f"  {row['template_code']}: {row['reusability_percentage']:.1f}% reusable")
        
        # Templates with most unique fields (customization needed)
        most_unique = df.nlargest(5, 'unique_to_this_doc')[['template_code', 'unique_to_this_doc']]
        logger.info(f"\n🎯 MOST UNIQUE TEMPLATES (Need Customization)")
        for _, row in most_unique.iterrows():
            logger.info(f"  {row['template_code']}: {row['unique_to_this_doc']} unique fields")


if __name__ == "__main__":
    reporter = ExcelTemplateReporter()
    
    # Analyze requirements first (self-reflection)
    reporter.analyze_requirements()
    
    # Get the main analysis
    main_df = reporter.get_matched_templates_with_full_metrics()
    
    # Print insights
    reporter.print_key_insights(main_df)
    
    # Generate the Excel report
    output_path = reporter.generate_excel_report()
    
    print(f"\n✅ Comprehensive report generated: {output_path}")
    print(f"   Analyzed {len(main_df)} matched Excel templates")